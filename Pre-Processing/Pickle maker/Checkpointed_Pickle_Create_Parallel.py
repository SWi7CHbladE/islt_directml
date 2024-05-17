import pickle
import gzip
import os
import shutil
import tempfile
import re
import cv2
import numpy as np
import torch
from concurrent.futures import ProcessPoolExecutor, as_completed
from tensorflow.keras.applications import EfficientNetB7
from tensorflow.keras.layers import GlobalAveragePooling2D
from tensorflow.keras.models import Model
from tensorflow.keras.applications.efficientnet import preprocess_input
from openpyxl import load_workbook
import portalocker
import datetime
import multiprocessing

def initialize_model():
    base_model = EfficientNetB7(weights='imagenet', include_top=False)
    x = base_model.output
    x = GlobalAveragePooling2D()(x)
    feature_extractor = Model(inputs=base_model.input, outputs=x)
    return feature_extractor

def save_checkpoint(checkpoint_path, checkpoint_name, list_of_inputs):
    unstable_path = os.path.join(checkpoint_path, "unstable")
    unstable_file = os.path.join(unstable_path, checkpoint_name)
    if not os.path.exists(unstable_path):
        os.makedirs(unstable_path)
    
    with tempfile.NamedTemporaryFile(delete=False) as tmp_file:
        with gzip.GzipFile(fileobj=tmp_file, mode='wb') as gz:
            pickle.dump(list_of_inputs, gz)
        tmp_filename = tmp_file.name

    try:
        shutil.move(tmp_filename, unstable_file)
        shutil.move(unstable_file, os.path.join(checkpoint_path, checkpoint_name))
        print("Backup made at: " + str(checkpoint_path))

        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        backup_folder = os.path.join(checkpoint_path, "backup")
        if not os.path.exists(backup_folder):
            os.makedirs(backup_folder)
        backup_unstable_folder = os.path.join(backup_folder, f"unstable_{checkpoint_name}_{timestamp}")
        shutil.copytree(unstable_path, backup_unstable_folder)
        print("Unstable folder backed up at: " + str(backup_unstable_folder))
    except Exception as e:
        print("\n\nError!!!! Could not create backup: " + str(e))
        exit()

def load_checkpoint(checkpoint_path, checkpoint_name):
    backup_file = os.path.join(checkpoint_path, checkpoint_name)
    if os.path.exists(backup_file):
        print("Loading from: " + str(backup_file))
        print("\n*************************\n*************************\n*************************\n*** Checkpoint Loaded ***\n*************************\n*************************\n*************************\n")
        try:
            with open(backup_file, 'rb') as f:
                portalocker.lock(f, portalocker.LOCK_SH)
                try:
                    with gzip.GzipFile(fileobj=f) as gz:
                        data = pickle.load(gz)
                finally:
                    portalocker.unlock(f)
                    print("File unlocked successfully")
            return data
        except Exception as e:
            print(f"Failed to read and lock the file: {e}")
            raise e
    else:
        print("Creating at: " + str(backup_file))
        print("\n****************************************\n****************************************\n****************************************\n*** Checkpoint Loading Failed!!!!!!! ***\n****************************************\n****************************************\n****************************************\n")
        return None

def get_features(filename, destination, feature_extractor):
    input_string = filename
    pattern = r'\d+'
    match = re.search(pattern, input_string)
    if match:
        first_match = match.group()
        input_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), destination, first_match, input_string)
        try:
            file_paths_frames = [file for file in sorted(os.listdir(input_folder)) if file.endswith(".jpg")]
        except:
            return None

        features_listofList = []
        for indx, frame_file in enumerate(file_paths_frames):
            frame_filename = os.path.join(input_folder, frame_file)
            image = cv2.imread(frame_filename)
            image = cv2.resize(image, (600, 600))
            image = preprocess_input(image)
            image = np.expand_dims(image, axis=0)
            spatial_embedding = feature_extractor.predict(image)[0]
            features_listofList.append(spatial_embedding)
        return torch.tensor(features_listofList)
    else:
        print("No match found for: " + input_string + "\n")
        return None

def create_pickle(config, frame_dest, checkpoint_path):
    vw_dest = config['vw_dest']
    vo_dest = config['vo_dest']
    checkpoint_name = config['checkpoint_name']

    feature_extractor = initialize_model()

    workbook = load_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)), vw_dest))
    sheet = workbook.active
    excel_data = []
    for row in sheet.iter_rows(values_only=True):
        excel_data.append(row)

    list_of_inputs = load_checkpoint(checkpoint_path, checkpoint_name)
    if list_of_inputs is None:
        list_of_inputs = []

    checkpoint_range = 20
    none_counter = 0
    flag = 0
    for index in range(len(list_of_inputs), len(excel_data), checkpoint_range):
        if flag == 1:
            exit()
        batch_list_of_inputs = []
        for tmp in excel_data[index:index + checkpoint_range]:
            features = get_features(str(tmp[0]), frame_dest, feature_extractor)
            if features is not None:
                none_counter = 0
                if len(features) > 0:
                    data_dict = {
                        'name': tmp[0],
                        'signer': tmp[1],
                        'gloss': tmp[2],
                        'text': tmp[3],
                        'sign': features + 1e-8
                    }
                    batch_list_of_inputs.append(data_dict)
            else:
                none_counter += 1
                if none_counter >= checkpoint_range - 1:
                    flag = 1
                    break
        if flag == 1:
            break
        
        list_of_inputs.extend(batch_list_of_inputs)

        save_checkpoint(checkpoint_path, checkpoint_name, list_of_inputs)
        torch.cuda.empty_cache()
        list_of_inputs = load_checkpoint(checkpoint_path, checkpoint_name)

    with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), vo_dest), 'wb') as f:
        with gzip.GzipFile(fileobj=f, mode='wb') as gz:
            pickle.dump(list_of_inputs, gz)

    print(f"Done processing {vw_dest}")

def run_multiple_pickle_creations(configurations, frame_dest, checkpoint_path):
    with ProcessPoolExecutor() as executor:
        futures = []
        for config in configurations:
            print(f"Submitting {config['vw_dest']} -> {config['vo_dest']} with checkpoint {config['checkpoint_name']}")
            futures.append(executor.submit(create_pickle, config, frame_dest, checkpoint_path))
        
        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                print(f"Error occurred: {e}")

if __name__ == "__main__":
    multiprocessing.freeze_support()

    # List of configurations to process
    configurations = [
        {
            'vw_dest': "Dataset/excels/Validation.xlsx",
            'vo_dest': "Dataset/Pickles/excel_data.dev",
            'checkpoint_name': 'dev_checkpoint.pkl'
        },
        {
            'vw_dest': "Dataset/excels/Test.xlsx",
            'vo_dest': "Dataset/Pickles/excel_data.test",
            'checkpoint_name': 'test_checkpoint.pkl'
        },
        {
            'vw_dest': "Dataset/excels/Train/Train_0.xlsx",
            'vo_dest': "Dataset/Pickles/excel_data0.train",
            'checkpoint_name': 'train_checkpoint_0.pkl'
        },
        {
            'vw_dest': "Dataset/excels/Train/Train_1.xlsx",
            'vo_dest': "Dataset/Pickles/excel_data1.train",
            'checkpoint_name': 'train_checkpoint_1.pkl'
        },
        {
            'vw_dest': "Dataset/excels/Train/Train_2.xlsx",
            'vo_dest': "Dataset/Pickles/excel_data2.train",
            'checkpoint_name': 'train_checkpoint_2.pkl'
        },
        {
            'vw_dest': "Dataset/excels/Train/Train_3.xlsx",
            'vo_dest': "Dataset/Pickles/excel_data3.train",
            'checkpoint_name': 'train_checkpoint_3.pkl'
        },
        {
            'vw_dest': "Dataset/excels/Train/Train_4.xlsx",
            'vo_dest': "Dataset/Pickles/excel_data4.train",
            'checkpoint_name': 'train_checkpoint_4.pkl'
        },
        {
            'vw_dest': "Dataset/excels/Train/Train_5.xlsx",
            'vo_dest': "Dataset/Pickles/excel_data5.train",
            'checkpoint_name': 'train_checkpoint_5.pkl'
        },
        {
            'vw_dest': "Dataset/excels/Train/Train_6.xlsx",
            'vo_dest': "Dataset/Pickles/excel_data6.train",
            'checkpoint_name': 'train_checkpoint_6.pkl'
        },
        {
            'vw_dest': "Dataset/excels/Train/Train_7.xlsx",
            'vo_dest': "Dataset/Pickles/excel_data7.train",
            'checkpoint_name': 'train_checkpoint_7.pkl'
        },
        {
            'vw_dest': "Dataset/excels/Train/Train_8.xlsx",
            'vo_dest': "Dataset/Pickles/excel_data8.train",
            'checkpoint_name': 'train_checkpoint_8.pkl'
        },
        {
            'vw_dest': "Dataset/excels/Train/Train_9.xlsx",
            'vo_dest': "Dataset/Pickles/excel_data9.train",
            'checkpoint_name': 'train_checkpoint_9.pkl'
        }
    ]

    vf_dest = "Dataset/Final folder for frames"
    store_to_path = 'C:\\Users\\Admin\\Rahul\\islt_directml\\Pre-Processing\\Pickle maker\\Dataset\\Checkpoint\\'

    run_multiple_pickle_creations(configurations, vf_dest, store_to_path)

    print("Done creating pickle files for all configurations.")





