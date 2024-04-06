import pickle
import gzip
import os
import shutil
from tensorflow import keras
import tensorflow as tf
import cv2
import numpy as np
import re
import os
import torch
from tqdm.notebook import tqdm
from tensorflow.keras.applications import imagenet_utils
from tensorflow.keras.applications.efficientnet import EfficientNetB7
from tensorflow.keras.layers import GlobalAveragePooling2D
from tensorflow.keras.models import Model
from tensorflow.keras.applications.efficientnet import preprocess_input
from tensorflow.keras import mixed_precision
from openpyxl import Workbook, load_workbook
import platform
from progress.bar import Bar

# Function to save checkpoint
def save_checkpoint(checkpoint_path, data, count):
    checkpoint_dir = os.path.dirname(checkpoint_path)
    if not os.path.exists(checkpoint_dir):
        os.makedirs(checkpoint_dir)
    with open(checkpoint_path, 'wb') as f:
        pickle.dump({'data': data, 'count': count}, f)

# Function to load checkpoint
def load_checkpoint(checkpoint_path):
    if os.path.exists(checkpoint_path):
        with open(checkpoint_path, 'rb') as f:
            return pickle.load(f)
    else:
        return None

# ***Function for extraction of features***
def get_features(filename, destination, checkpoint_path):
    checkpoint_data = load_checkpoint(checkpoint_path)
    checkpoint_count = checkpoint_data['count'] if checkpoint_data else 0
    processed_data = checkpoint_data['data'] if checkpoint_data else []
    
    input_string = filename
    pattern = r'\d+'
    match = re.search(pattern, input_string)
    
    if match:
        first_match = match.group()
        input_folder = os.path.join(os.getcwd(), destination, first_match, input_string)
        
        try:
            file_paths_frames = [file for file in sorted(os.listdir(input_folder)) if file.endswith(".jpg")]
        except:
            return None

        base_model = EfficientNetB7(weights='EfficientNet7_Emot.h5', include_top=False)
        x = base_model.output
        x = GlobalAveragePooling2D()(x)
        feature_extractor = Model(inputs=base_model.input, outputs=x)
    
        features_listofList = []
        for indx, frame_file in enumerate(file_paths_frames):
            frame_filename = os.path.join(input_folder, frame_file)
            image = cv2.imread(frame_filename)
            image = cv2.resize(image, (600, 600))
            image = preprocess_input(image)
            image = np.expand_dims(image, axis=0)
            spatial_embedding = feature_extractor.predict(image)[0]
            features_listofList.append(spatial_embedding)
            checkpoint_count += 1
            processed_data.append({'filename': frame_file, 'features': spatial_embedding})
            if checkpoint_count % 10 == 0:
                print("Checkpoint created:", checkpoint_count)
                save_checkpoint(checkpoint_path, processed_data, checkpoint_count)

        save_checkpoint(checkpoint_path, processed_data, checkpoint_count)
        return torch.tensor(features_listofList)
    else:
        print("No match found.")
        return None

# Function to create the pickle file
def create_pickle(workbook_dest, output_dest, frame_dest, checkpoint_path):
    workbook = load_workbook(workbook_dest)
    sheet = workbook.active
    excel_data = []
    for row in sheet.iter_rows(values_only=True):
        excel_data.append(row)
    
    list_of_inputs = []
    for tmp in excel_data:
        features = get_features(str(tmp[0]), frame_dest, checkpoint_path)
        if features is not None:
            if len(features) > 0:
                data_dict = {
                    'name': tmp[0],
                    'signer': tmp[1],
                    'gloss': tmp[2],
                    'text': tmp[3],
                    'sign': features + 1e-8
                }
                list_of_inputs.append(data_dict)
    
    with gzip.open(os.path.join(os.getcwd(), output_dest), 'wb') as f:
        pickle.dump(list_of_inputs, f)

# Files to access
vw_dest = "Dataset/excels/Validation.xlsx"
vo_dest = "Dataset/Pickles/excel_data.dev"
vf_dest = "Dataset/Final folder for frames"

tw_dest = "Dataset/excels/Test.xlsx"
to_dest = "Dataset/Pickles/excel_data.test"
tf_dest = "Dataset/Final folder for frames"

w_dest = "Dataset/excels/Train.xlsx"
o_dest = "Dataset/Pickles/excel_data.train"
f_dest = "Dataset/Final folder for frames"

# Checkpoint paths
dev_checkpoint_path = 'Dataset/dev_checkpoint.pkl'
test_checkpoint_path = 'Dataset/test_checkpoint.pkl'
train_checkpoint_path = 'Dataset/train_checkpoint.pkl'

# Create pickles
create_pickle(vw_dest, vo_dest, vf_dest, dev_checkpoint_path)
create_pickle(tw_dest, to_dest, tf_dest, test_checkpoint_path)
create_pickle(w_dest, o_dest, f_dest, train_checkpoint_path)

print("Done creating pickle files.")
