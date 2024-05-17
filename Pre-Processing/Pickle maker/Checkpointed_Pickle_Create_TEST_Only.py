import pickle
import gzip
import os
import shutil
from tensorflow import keras
import tensorflow as tf
import cv2
import numpy as np
import re
import os
import torch
from tqdm.notebook import tqdm
from tensorflow.keras.applications import imagenet_utils
from tensorflow.keras.applications.efficientnet import EfficientNetB7
from tensorflow.keras.layers import GlobalAveragePooling2D
from tensorflow.keras.models import Model
from tensorflow.keras.applications.efficientnet import preprocess_input
from tensorflow.keras import mixed_precision
from openpyxl import Workbook, load_workbook
import platform
from progress.bar import Bar
import shutil

# initialise the model
base_model = EfficientNetB7(weights='imagenet', include_top=False)
x = base_model.output
x = GlobalAveragePooling2D()(x)
feature_extractor = Model(inputs=base_model.input, outputs=x)


print("***\nCurrent working directory:\n")
print(os.getcwd())
print("***")
# Function to save checkpoint
def save_checkpoint(checkpoint_path, checkpoint_name, list_of_inputs):
    unstable_path = checkpoint_path + r"unstable\\"
    unstable_file = unstable_path + checkpoint_name
    #print("saving at: "+ str(os.path.join(os.path.dirname(os.path.abspath(__file__)), checkpoint_path)))
    #print("saving at: "+ str(checkpoint_path))
    #with gzip.open(os.path.join(os.path.dirname(os.path.abspath(__file__)), checkpoint_path), 'wb') as f:
    with gzip.open(unstable_file, 'wb') as f:
        pickle.dump(list_of_inputs, f)
        print("\n************************\n************************\n************************\n*** Checkpoint Saved ***\n************************\n************************\n************************\n")
    if not os.path.exists(unstable_path):
        os.makedirs(unstable_path)
    try:
        shutil.move(unstable_file, checkpoint_path)
        print("backup made at: "+str(checkpoint_path))
    except:
        print("\n\nError!!!! Could not create backup")
        exit()

# Function to load checkpoint
def load_checkpoint(checkpoint_path, checkpoint_name):
    backup_file = checkpoint_path + checkpoint_name
    #checkpoint_path  = os.path.join(os.path.dirname(os.path.abspath(__file__)), checkpoint_path)
    if os.path.exists(backup_file):
        print("loading from: "+ str(backup_file))
        print("\n*************************\n*************************\n*************************\n*** Checkpoint Loaded ***\n*************************\n*************************\n*************************\n")
        with gzip.open(backup_file, 'rb') as f:
            return pickle.load(f)
    else:
        print("creating at: "+ str(backup_file))
        print("\n****************************************\n****************************************\n****************************************\n*** Checkpoint Loading Failed!!!!!!! ***\n****************************************\n****************************************\n****************************************\n")
        return None

# Function for extraction of features
def get_features(filename, destination):
    input_string = filename
    pattern = r'\d+'
    match = re.search(pattern, input_string)
    if match:
        first_match = match.group()
        input_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), destination, first_match, input_string)
        try:
            file_paths_frames = [file for file in sorted(os.listdir(input_folder)) if file.endswith(".jpg")]
        except:
            return None

        # base_model = EfficientNetB7(weights='imagenet', include_top=False)
        # x = base_model.output
        # x = GlobalAveragePooling2D()(x)
        # feature_extractor = Model(inputs=base_model.input, outputs=x)

        features_listofList = []
        for indx, frame_file in enumerate(file_paths_frames):
            frame_filename = os.path.join(input_folder, frame_file)
            image = cv2.imread(frame_filename)
            image = cv2.resize(image, (600, 600))
            image = preprocess_input(image)
            image = np.expand_dims(image, axis=0)
            spatial_embedding = feature_extractor.predict(image)[0]
            features_listofList.append(spatial_embedding)
        return torch.tensor(features_listofList)
    else:
        print("No match found for: " + input_string + "\n")
        return None

# Function to create the pickle file
def create_pickle(workbook_dest, output_dest, frame_dest, checkpoint_path, filename):
    workbook = load_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)),workbook_dest))
    sheet = workbook.active
    excel_data = []
    for row in sheet.iter_rows(values_only=True):
        excel_data.append(row)

    # Load checkpoint
    list_of_inputs = load_checkpoint(checkpoint_path, filename)
    if list_of_inputs is None:
        list_of_inputs = []

    # Get the features
    checkpoint_range = 20
    none_counter = 0
    flag = 0
    for index in range(len(list_of_inputs), len(excel_data), checkpoint_range):
        if flag == 1:
            exit()
        batch_list_of_inputs = []
        for tmp in excel_data[index:index + checkpoint_range]:
            features = get_features(str(tmp[0]), frame_dest)
            if features is not None:
                none_counter = 0
                if len(features) > 0:
                    data_dict = {
                        'name': tmp[0],
                        'signer': tmp[1],
                        'gloss': tmp[2],
                        'text': tmp[3],
                        'sign': features + 1e-8
                    }
                    batch_list_of_inputs.append(data_dict)
            else:
                none_counter += 1
                if(none_counter >= checkpoint_range - 1):
                    flag = 1
                    break
        if flag == 1:
            break
        
        # Update list_of_inputs
        list_of_inputs.extend(batch_list_of_inputs)

        # Save checkpoint
        save_checkpoint(checkpoint_path, filename, list_of_inputs)
        torch.cuda.empty_cache()
        list_of_inputs = load_checkpoint(checkpoint_path, filename)


    # Save final pickle file
    with gzip.open(os.path.join(os.path.dirname(os.path.abspath(__file__)), output_dest), 'wb') as f:
        pickle.dump(list_of_inputs, f)


# Files to access
tw_dest = "Dataset/excels/Test.xlsx"
to_dest = "Dataset/Pickles/excel_data.test"
tf_dest = "Dataset/Final folder for frames"
checkpoint_name = 'test_checkpoint.pkl'
store_to_path = 'C:\\Users\\Admin\\Rahul\\islt_directml\\Pre-Processing\\Pickle maker\\Dataset\\Checkpoint\\'

create_pickle(tw_dest, to_dest, tf_dest, store_to_path, checkpoint_name)




print("Done creating pickle files.")
