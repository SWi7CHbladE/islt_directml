import pickle
import gzip
import os
from xml.etree.ElementTree import tostring
from tensorflow import keras
import tensorflow as tf
from tensorflow.keras import mixed_precision
# import tensorflow_addons as tfa
# import keras_tuner
import cv2
import numpy as np
import re
import os
import torch
from tqdm.notebook import tqdm
from tensorflow.keras.applications import imagenet_utils
from tensorflow.keras.applications.efficientnet import EfficientNetB7, EfficientNetB0 
from tensorflow.keras.layers import GlobalAveragePooling2D
from tensorflow.keras.models import Model
from tensorflow.keras.applications.efficientnet import preprocess_input
from tensorflow.keras import mixed_precision
from openpyxl import Workbook, load_workbook
import platform
from progress.bar import Bar
import tqdm
import multiprocessing
import concurrent.futures


tf.config.list_physical_devices('GPU')
tf.keras.backend.clear_session()
# tf.config.run_functions_eagerly(True)


## quietly deep-reload tqdm
# import sys
# from IPython.lib import deepreload 
# stdout = sys.stdout
# sys.stdout = open('junk','w')
# deepreload.reload(tqdm)
# sys.stdout = stdout

# tqdm.__version__

## Training using FP16
# mixed_precision.global_policy()
# mixed_precision.set_global_policy('mixed_float16')
# mixed_precision.global_policy()


def cls():
    system = platform.system()
    if system == 'Windows':
        os.system('cls')
    elif system == 'Linux':
        os.system('clear')
    elif system == 'MacOS':
        os.system('clear')




# ***Initialise the CNN model***
base_model = EfficientNetB7(weights='imagenet', include_top=False)
x = base_model.output
x = GlobalAveragePooling2D()(x)
feature_extractor = Model(inputs=base_model.input, outputs=x)




# ***Function for extraction of features***
def get_features(filename, destination):
    input_string = filename
    pattern = r'\d+'
    match = re.search(pattern, input_string)
    features_listofList=[]
    if match:
        first_match = match.group()
        #print(first_match)
        # Get file path name
        input_folder = os.getcwd()+"/"+destination+"/"+first_match+"/"+input_string
        try:
            file_paths_frames = [file for file in sorted(os.listdir(input_folder)) if file.endswith(".jpg")]

            # base_model needs to be initialised first
            # x = base_model.output
            # x = GlobalAveragePooling2D()(x)
            # feature_extractor = Model(inputs=base_model.input, outputs=x)
    
            
            for indx, frame_file in enumerate(file_paths_frames):
                # Get the file
                frame_filename =  input_folder+"/" + frame_file
                print("frame_filename============", frame_filename,"\n")
                image = cv2.imread(frame_filename)
                
                #Image processing part
        
                # Define the coordinates of the ROI (top-left and bottom-right)
                x1, y1 = 535, 0  # Top-left corner of ROI
                x2, y2 = 1385, 1080  # Bottom-right corner of ROI
                #210*260
                # Crop the image to the specified ROI
                image = image[y1:y2, x1:x2]
                image = cv2.resize(image, (224, 224))  # Resize to the input size of MobileNet

                # Preprocess the image
                image = preprocess_input(image)
        
                # Expand dimensions to match the expected input shape (batch size of 1)
                image = np.expand_dims(image, axis=0)
                
                # Extract spatial embedding
                spatial_embedding = feature_extractor.predict(image)[0]
                
                features_listofList.append(spatial_embedding)
        except:
            print("The file is missing in the frames dataset\n")
    else:
        print("No match found.\n")
        return None

    # after this, the video must be in the form of features
    return torch.tensor(np.array(features_listofList))
    # return torch.tensor(features_listofList)




# ***Function to create the pickle file***
def create_pickle(workbook_dest, output_dest, frame_dest):
    #load the excel file
    workbook = load_workbook(workbook_dest)
    sheet = workbook.active

    # Extract data from the Excel file
    excel_data = []
    for row in sheet.iter_rows(values_only=True):
        excel_data.append(row)
        # print(excel_data,"\n")
        # print(excel_data)list_of_inputs = []
    
    # Get the features
    list_of_inputs = []
    
    for tmp in excel_data:
        features = get_features(str(tmp[0]),frame_dest)
        if(features!= None):
            data_dict = {}
            data_dict['name'] = tmp[0]
            data_dict['signer'] = tmp[1]
            data_dict['gloss'] = tmp[2]
            data_dict['text'] = tmp[3]
            data_dict['sign'] = features + 1e-8
            #print(data_dict)
            #input()
            list_of_inputs.append(data_dict)
        
    # print("\nlist_of_input:\n")
    # print(list_of_inputs)
    with gzip.open(os.getcwd() + "/" + output_dest,'wb') as f:
        pickle.dump(list_of_inputs,f)



# Files to access:
# validation
vw_dest = "Dataset/excels/Sports_dataset"
vo_dest = "Dataset/Pickles/Sports_dataset"
vf_dest = "Dataset/Final folder for frames"

# test
tw_dest = "Dataset/excels/Sports_dataset"
to_dest = "Dataset/Pickles/Sports_dataset"
tf_dest = "Dataset/Final folder for frames"

# train
w_dest = "Dataset/excels/Sports_dataset"
o_dest = "Dataset/Pickles/Sports_dataset"
f_dest = "Dataset/Final folder for frames"

val_excel_postfix = "ISH_News_Sports final VAL.xlsx"
test_excel_postfix = "ISH_News_Sports final TEST.xlsx"
train_excel_postfix = "ISH_News_Sports final TRAIN.xlsx"

val_output_postfix = "excel_data.dev"
test_output_postfix = "excel_data.test"
train_output_postfix = "excel_data.train"



i = 0

while i < 10:
    # Validation pickle
    tmp1 = vw_dest + "/" + str(i) + "/" + val_excel_postfix
    tmp2 = vo_dest + "/" + str(i) + "/" + val_output_postfix
    create_pickle(tmp1, tmp2, vf_dest)


    # Testing pickle
    tmp1 = tw_dest + "/" + str(i) + "/" + test_excel_postfix
    tmp2 = to_dest + "/" + str(i) + "/" + test_output_postfix
    create_pickle(tmp1, tmp2, tf_dest)


    # Training pickle
    tmp1 = w_dest + "/" + str(i) + "/" + train_excel_postfix
    tmp2 = o_dest + "/" + str(i) + "/" + train_output_postfix
    create_pickle(tmp1, tmp2, f_dest)
    
    i+=1


print("Done")